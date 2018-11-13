import requests,bs4,openpyxl,re
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def writeExcel(names,data):
    wb = openpyxl.Workbook()
    sheet=wb.active
    columns= ["Name","Interests","Link","Facebook","Twitter","Linkedin",
                "Email","Phone","Address"]
    # vars = [names,interest_lists,websites,data["Facebook"],data["Twitter"],
    #         data["Linkedin"],data["Email"],data["Phone"],data['Address']]
    for i in range(len(columns)):
        sheet.cell(row=1,column=i+1).value = columns[i]
        sheet.cell(row=1,column=i+1).fill = PatternFill(bgColor="0735b6", fill_type = "solid")
        sheet.cell(row=1,column=i+1).font.color.index = Color.WHITE

    dims={"Name":[],"Interests":[],"Link":[],"Facebook":[],
            "Twitter":[],"Linkedin":[],"Email":[],"Phone":[],
            "Address":[]}
    for i in range(len(data)):
        for j in range(len(columns)):
            sheet.cell(row=i+2,column=j+1).value ='\n'.join(data[names[i]][columns[j]])
            dims[columns[j]].append(len('\n'.join(data[names[i]][columns[j]])))

    for i in range(len(columns)):
        sheet.column_dimensions[get_column_letter(i+1)].width = max(dims[columns[i]])

    wb.freeze_panes=sheet['A2']
    # for column_cells in worksheet.columns:
    # length = max(len(as_text(cell.value)) for cell in column_cells)
    # worksheet.column_dimensions[column_cells[0].column].width = length
    wb.save('driven_AI3.xlsx')


def getInfo(link):
    address_list=[]
    phone_list=[]
    email_list=[]
    facebook_links=[]
    twitter_links=[]
    linkedin_links=[]
    linkOptions = ['contact-us','contact','about-us','about','']
    for option in linkOptions:
        # print(link+option)

        try:
            res= requests.get(link+option)
            res.raise_for_status()
        except:continue
        soup = bs4.BeautifulSoup(res.text,'lxml')
        addressReg = re.compile(r'[^<>{}:\"], Cairo\W?$|[^<>{}:\"], Egypt\W?$',re.IGNORECASE)
        phoneReg = re.compile(r'\+\W?2\W?\d?\d?\W?\d{4}\W?\d{4}|\+\W?02\W?\d?\W?\d{4}\W?\d{4}')
        emailReg=re.compile(r'[^:\"<>/\s,][A-Za-z0-9._-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,4}$',re.MULTILINE)
        address_list = [*{e.strip('\n<> ').split(":")[-1] for e in soup(text=addressReg)}]
        # print(address_list)
        phone_list = [*{e.strip('\n<> ').split('+')[-1] for e in soup(text=phoneReg)}]
        email_list = [*{e.strip('\n<> ') for e in soup(text=emailReg)}]

        facebook_links = [*{x.get('href').lower().split("//")[1] for x in soup.select('a')\
                            if "facebook.com" in x.get('href',[])}]
        twitter_links = [*{x.get('href').lower().split("//")[1] for x in soup.select('a')\
                            if "twitter.com" in x.get('href',[])}]
        linkedin_links = [*{x.get('href').lower().split("//")[1] for x in soup.select('a')\
                            if "linkedin.com" in x.get('href',[])}]

        if address_list:
            break
    return({"Address":address_list,"Phone":phone_list,"Email":email_list,\
                            "Facebook":facebook_links,"Twitter":twitter_links,"Linkedin":linkedin_links})

res = requests.get("https://my-interviews-experience-in-egypt.quora.com/AI-Driven-Companies-in-Egypt")
res.raise_for_status()
soup=bs4.BeautifulSoup(res.text,'lxml')

pageElem = soup.select('ol')
companyElem = pageElem[0].select('li')

names = []

data=dict()
for company in companyElem:
    name= company.text.split(':')[0]
    website=company.select('a')[0].get('href')
    company_interests = company.text.split(':')[-1]
    interest_list =company_interests.split('=')[0]

    temp=getInfo(website)
    temp['Name']=name
    temp['Link']=website
    temp['Interests']=interest_list
    data[name]=temp
    names.append(name)

writeExcel(names,data)

# # print(getInfo("http://www.cognitev.com/"))
# # print(getInfo("http://www.itworx.com/"))
# # print(getInfo("https://stratochem.com/"))
# print(getInfo("https://yaoota.com/en-eg/"))
