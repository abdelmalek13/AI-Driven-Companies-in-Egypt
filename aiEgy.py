#! python3
import requests,bs4,openpyxl,re
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill,Font

def writeExcel(names,data):
    ## write the resulted data in a new excelsheet and save it with name driven_AI3.xlsx
    ## input: names --a list of names of the companies
    ##        data -- all the data of the companies in the form of dict that includes
    ##                name of companies, thier fields of interest, website link,
    ##                link to facebook page, link to twitter account, link to linkedin account,
    ##                email/s of the company has, thier phone numbers, and thier address in Egypt

    wb = openpyxl.Workbook()
    sheet=wb.active
    columns= ["Name","Interests","Link","Facebook","Twitter","Linkedin",
                "Email","Phone","Address"]

    for i in range(len(columns)):
        # add header column with the names of the columns and format it
        sheet.cell(row=1,column=i+1).value = columns[i]
        sheet.cell(row=1,column=i+1).fill = PatternFill(bgColor="000000FF", fill_type = "solid")
        sheet.cell(row=1,column=i+1).font = Font(color="00FFFFFF")

    dims={"Name":[],"Interests":[],"Link":[],"Facebook":[],
            "Twitter":[],"Linkedin":[],"Email":[],"Phone":[],
            "Address":[]}
    for i in range(len(data)):
        for j in range(len(columns)):
            # write the data dict to the excel sheet
            sheet.cell(row=i+2,column=j+1).value ='\n'.join(data[names[i]][columns[j]]) #in case that a company has two results in certain column, joins them with newline
            dims[columns[j]].append(len('\n'.join(data[names[i]][columns[j]]))) #store the length of the string in order to be used in modifying the width of columns

    for i in range(len(columns)):
        #Modify the width of the column using the maximum length of string in each column
        sheet.column_dimensions[get_column_letter(i+1)].width = max(dims[columns[i]])

    sheet.freeze_panes=sheet['B2']
    wb.save('driven_AI3.xlsx')


def getInfo(link):
    ## Get all the data of a company through it's website
    ## input: link -- the company's website link address
    ## output: a dictionary includes:
    ##            link to facebook page, link to twitter account, link to linkedin account,
    ##            email/s of the company has, thier phone numbers, and thier address in Egypt
    address_list=[]
    phone_list=[]
    email_list=[]
    facebook_links=[]
    twitter_links=[]
    linkedin_links=[]
    linkOptions = ['contact-us','contact','about-us','about','']

    for option in linkOptions:
        # Try every page that can contains the contact info from 5 possible page links
        try:
            res= requests.get(link+option)
            res.raise_for_status()
        except:continue
        soup = bs4.BeautifulSoup(res.text,'lxml')
        addressReg = re.compile(r'[^<>{}:\"], Cairo\W?$|[^<>{}:\"], Egypt\W?$',re.IGNORECASE)
        phoneReg = re.compile(r'\+\W?2\W?\d?\d?\W?\d{4}\W?\d{4}|\+\W?02\W?\d?\W?\d{4}\W?\d{4}')
        emailReg=re.compile(r'[^:\"<>/\s,][A-Za-z0-9._-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,4}$',re.MULTILINE)
        address_list = [*{e.strip('\n<> ').split(":")[-1] for e in soup(text=addressReg)}] #locate the address , ex--> Address: Maadi Technology Park, 11435 Cairo
        phone_list = [*{e.strip('\n<> ').split('+')[-1] for e in soup(text=phoneReg)}] #locate the phone number , ex--> +2 02 23456789
        email_list = [*{e.strip('\n<> ') for e in soup(text=emailReg)}] #locate the email, ex--> <contact@co.com>

        facebook_links = [*{x.get('href').lower().split("//")[1] for x in soup.select('a')\
                            if "facebook.com" in x.get('href',[])}] #get the Facebook link from all the links in contact page and remove "http://"
        twitter_links = [*{x.get('href').lower().split("//")[1] for x in soup.select('a')\
                            if "twitter.com" in x.get('href',[])}] #get the Twitter link from all the links in contact page and remove "http://"
        linkedin_links = [*{x.get('href').lower().split("//")[1] for x in soup.select('a')\
                            if "linkedin.com" in x.get('href',[])}] #get the Linkedin link from all the links in contact page and remove "http://"

        if address_list:
            # stop in case the crawler already found the data
            break
    return({"Address":address_list,"Phone":phone_list,"Email":email_list,\
                            "Facebook":facebook_links,"Twitter":twitter_links,"Linkedin":linkedin_links})

if __name__ =="__main__":
    res = requests.get("https://my-interviews-experience-in-egypt.quora.com/AI-Driven-Companies-in-Egypt")
    res.raise_for_status()
    soup=bs4.BeautifulSoup(res.text,'lxml')

    pageElem = soup.select('ol')
    companyElem = pageElem[0].select('li')

    names = []

    data=dict()
    for company in companyElem:
        ## get the data for every company, and store it in dict with its name as a key
        name= company.text.split(':')[0]
        website=company.select('a')[0].get('href')
        company_interests = company.text.split(':')[-1]
        interest_list =company_interests.split('=')[0]  #get rid of "======" in the end of each company section

        temp=getInfo(website)
        temp['Name']=name
        temp['Link']=website
        temp['Interests']=interest_list
        data[name]=temp
        names.append(name)

    writeExcel(names,data)
